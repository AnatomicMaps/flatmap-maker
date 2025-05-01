#===============================================================================
#
#  Flatmap viewer and annotation tools
#
#  Copyright (c) 2019 - 2023 David Brooks
#
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#
#===============================================================================

from datetime import datetime
from pathlib import Path
from dataclasses import dataclass
import importlib
from io import BytesIO
import json
import logging
import mimetypes
import os
from pathlib import Path
import shutil
from zipfile import ZipFile, ZipInfo, ZIP_DEFLATED
from typing import Optional

#===============================================================================

import openpyxl
import requests

#===============================================================================

from mapmaker.flatmap import FlatMap

#===============================================================================

MAPPING_DATA_JSON = str(importlib.resources.files('mapmaker/output/data_mapping.json'))

#===============================================================================

from mapmaker.utils import pathlib_path

#===============================================================================

class VersionMapping:
    def __init__(self):
        with open(MAPPING_DATA_JSON, 'r') as f:
            self.__mappings = json.load(f)

    @property
    def available_versions(self):
        return [v['version'] for v in self.__mappings]

    def get_mapping(self, other_params):
        """
        : other_params: is a dictionary containing other data such as uuid and version
        """
        version = other_params.get('version', None)
        mapping = None
        if version == None:
            mapping = self.__mappings[0]
        else:
            for v in self.__mappings:
                if v['version'] == version:
                    mapping = v
        if mapping == None:
            raise Exception('Dataset-Description version-{} is not available'.format(version))
        for m in mapping['mapping']:
            if len(m[1])> 0:
                param = m[1][-1]
                if param in other_params:
                    m[2] = other_params[param]
        return mapping

#===============================================================================

class DatasetDescription:
    def __init__(self, flatmap, version):
        """
        : description_file: is a pat to description.json
        : flatmap: is a Flatmap instance
        : version: is SDS version
        """
        
        other_params = {
            'version': version,
            'id': [flatmap.metadata.get('source'), flatmap.uuid],
            'id_type': ['SourceManifestUrl', 'UUID'],
        }
        self.__mapping = VersionMapping().get_mapping(other_params)
        self.__workbook = self.__load_template_workbook(self.__mapping['template_url'])
        
    def write(self, description_file):
        if description_file.startswith('file'):
            description_file = pathlib_path(description_file)
        with open(description_file, 'r') as fd:
            self.__description = json.load(fd)
        for m in self.__mapping['mapping']:
             self.__write_cell(m)
        
    def __load_template_workbook(self, template_link):
        """
        : template_link: link to dataset_description.xlsx
        """
        headers = {'Content-Type': 'application/xlsx'}
        template = requests.request('GET', template_link, headers=headers)
        workbook = openpyxl.load_workbook(BytesIO(template.content))
        return workbook
        
    def __write_cell(self, map):
        worksheet = self.__workbook.worksheets[0]
        data_pos = self.__mapping.get('data_pos', 3)
        key, dsc, default = map
        values = default if isinstance(default, list) else [default]

        if len(dsc) == 1:
            if dsc[0] in self.__description:
                values = self.__description[dsc[-1]] if isinstance(self.__description[dsc[-1]], list) else [self.__description[dsc[-1]]]
        elif len(dsc) > 1:
            tmp_values = self.__description
            for d in dsc:
                if isinstance(tmp_values, dict):
                    tmp_values = tmp_values.get(d, {})
                elif isinstance(tmp_values, list):
                    tmp_values = [val.get(d, '') for val in tmp_values]
            
            if len(tmp_values) > 0:
                values = tmp_values if isinstance(tmp_values, list) else [tmp_values]

        for row in worksheet.rows:
            if row[0].value == None:
                break
            if row[0].value.lower().strip() == key:
                for pos in range(len(values)):
                    row[pos+data_pos].value = str(values[pos])

    def get_bytes(self):
        buffer = BytesIO()
        self.__workbook.save(buffer)
        buffer.seek(0)
        return buffer
    
    def get_json(self):
        return self.__description
    
    def close(self):
        self.__workbook.close()

#===============================================================================

@dataclass
class DatasetFile:
    filename: str
    fullpath: Path
    timestamp: datetime
    description: str
    file_type: str

#===============================================================================

class DirectoryManifest:
    COLUMNS = (
        'filename',
        'timestamp',
        'description',
        'file type',
    )

    def __init__(self, metadata: Optional[dict[str, str]]=None):
        self.__metadata = metadata if metadata is not None else {}
        self.__files = []
        self.__file_records = []

    @property
    def files(self):
        return self.__files
    
    @property
    def file_list(self):
        return [f.fullpath for f in self.__files]

    def add_file(self, filepath, description, timestamp):
        fullpath = Path(filepath)
        if fullpath.parts[0] == 'file:':
            fullpath = Path('/', *fullpath.parts[1:])
        file_type = mimetypes.guess_type(fullpath, strict=False)[0]
        file_type = fullpath.suffix if file_type == None else file_type
        dataset_file = DatasetFile(fullpath.name,
                                   fullpath,
                                   timestamp,
                                   description,
                                   file_type)
        self.__files.append(dataset_file)
        record: list[str|None] = [
            dataset_file.filename,
            dataset_file.timestamp.isoformat(timespec='milliseconds'),
            dataset_file.description,
            dataset_file.file_type
        ]
        for value in self.__metadata.values():
            record.append(value)
        self.__file_records.append(record)

    def __get_bytes(self):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        for col, value in enumerate(self.COLUMNS + tuple(self.__metadata.keys()), start=1):
            worksheet.cell(row=1, column=col, value=value)
        for row, record in enumerate(self.__file_records, start=2):
            for col, value in enumerate(record, start=1):
                worksheet.cell(row=row, column=col, value=value)
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        workbook.close()
        return buffer

    def copy_to_archive(self, archive: ZipFile, target: str):
        for file in self.files:
            zinfo = ZipInfo.from_file(file.fullpath, arcname=f'{target}/{file.filename}')
            zinfo.compress_type = ZIP_DEFLATED
            timestamp = file.timestamp
            zinfo.date_time = (timestamp.year, timestamp.month, timestamp.day,
                               timestamp.hour, timestamp.minute, timestamp.second)
            with open(file.fullpath, "rb") as src, archive.open(zinfo, 'w') as dest:
                shutil.copyfileobj(src, dest, 1024*8)
        manifest_workbook = self.__get_bytes()
        archive.writestr(f'{target}/manifest.xlsx', manifest_workbook.getvalue())
        manifest_workbook.close()

#===============================================================================

class FlatmapDirectoryManifests:
    def __init__(self, flatmap: FlatMap):
        # creating dataset manifest.xlsx

        manifest = flatmap.manifest
        species = manifest.models
        metadata = {'species': species} if species is not None else {}

        self.__dataset_image: Optional[Path] = None
        # Scan the output directory first since this is where we expect an SVG to use for the image
        self.__dataset_image = None
        self.__derivative_manifest = DirectoryManifest(metadata)
        for file in os.listdir(flatmap.map_dir):
            fullpath = (Path(flatmap.map_dir) / file).resolve()
            if fullpath.is_file():
                self.__derivative_manifest.add_file(fullpath, 'Generarate file for flatmap server',
                    datetime.fromtimestamp(os.path.getmtime(fullpath)))
                if self.__dataset_image is None and fullpath.suffix == '.svg':
                    self.__dataset_image = fullpath

        self.__primary_manifest = DirectoryManifest(metadata)
        git_status = manifest.git_status
        if git_status is None:
            raise TypeError('Map sources must be all be committed into git before SDS creation')
        source_timestamp: datetime = git_status['committed']
        for manifest_file in manifest.file_set:
            fullpath = Path(manifest_file.path)
            self.__primary_manifest.add_file(fullpath, manifest_file.description, source_timestamp)
            if self.__dataset_image is None and fullpath.suffix == '.svg':
                self.__dataset_image = fullpath

    @property
    def dataset_image(self):
        return self.__dataset_image

    def copy_to_archive(self, archive: ZipFile):
        self.__primary_manifest.copy_to_archive(archive, 'files/primary')
        self.__derivative_manifest.copy_to_archive(archive, 'files/derivative')

#===============================================================================

class SparcDataset:
    def __init__(self, flatmap: FlatMap):
        self.__flatmap = flatmap
        
    def generate(self):
        # generate dataset_description
        self.__description = DatasetDescription(self.__flatmap, version=None)
        try:
            self.__description.write(self.__flatmap.manifest.description)
        except:
            logging.error(f'Cannot create dataset: Cannot open: {self.__flatmap.__manifest.description}')

        # generate flatmap directory manifests
        self.__flatmap_manifests = FlatmapDirectoryManifests(self.__flatmap)

    def save(self, dataset: str):
        # create archive
        dataset_archive = ZipFile(dataset, mode='w', compression=ZIP_DEFLATED)

        # adding dataset_description
        desc_bytes = self.__description.get_bytes()
        dataset_archive.writestr('files/dataset_description.xlsx', desc_bytes.getvalue())
        desc_bytes.close()
        self.__description.close()
        
        # copy data
        self.__flatmap_manifests.copy_to_archive(dataset_archive)

        # create and save proper readme file, generated for dataset_description
        self.__add_readme(dataset_archive)

        # save banner
        banner_file = self.__flatmap_manifests.dataset_image
        if banner_file is not None:
            dataset_archive.write(banner_file, 'files/banner.svg')

        # close archive
        dataset_archive.close()

    def __add_readme(self, archive):
        # load flatmap description
        readme = ['# FLATMAP DESCRIPTION'] + self.__metadata_parser(self.__description.get_json())
        # load flatmap setup
        readme += ['# FLATMAP SETTINGS'] + self.__metadata_parser(self.__flatmap.metadata)
        archive.writestr(f'files/readme.md', '\n'.join(readme))

    def __metadata_parser(self, data):
        metadata = []
        for key, val in data.items():
                metadata += [f'## {key.capitalize()}']
                if isinstance(val, dict):
                    for subkey, subval in val.items():
                        metadata += [f'- {subkey}: {subval}']
                elif isinstance(val, list):
                    for subval in val:
                        if isinstance(subval, dict):
                            for subsubkey, subsubval in subval.items():
                                metadata += [f'- {subsubkey}: {subsubval}']
                            metadata += ['\n', '<br/>', '\n']
                        else:
                            metadata += [f'- {subval}']
                else:
                    metadata += [str(val), '\n']
        return metadata
        
#===============================================================================
